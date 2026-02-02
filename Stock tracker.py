from openpyxl import load_workbook
from openpyxl.styles import Font
from bs4 import BeautifulSoup
from datetime import date
from subprocess import Popen
from customtkinter import *
from PIL import Image
from time import sleep
from tkinter import messagebox
import threading
import yfinance as yf
import os

# --- CONFIGURAÇÕES DE CAMINHO ---
planilha = "planilha_acoes.xlsx"

# --- FUNÇÕES DE BACKEND ---

def obter_dados(ticker):
    try:
        ticker_sa = f"{ticker.upper()}.SA"
        acao = yf.Ticker(ticker_sa)
        info = acao.info
        cotacao = info.get('currentPrice', 0.0)
        dy = info.get('dividendYield', 0.0) * 100
        pvp = info.get('priceToBook', 0.0)
        pl = info.get('forwardPE', 0.0)
        div_ebitda = info.get('enterpriseToEbitda', 0.0)
        
        return (f"{cotacao:.2f}", f"{dy:.2f}%", f"{pvp:.2f}", f"{pl:.2f}", f"{div_ebitda:.2f}")
    except:
        return 'N/A', 'N/A', 'N/A', 'N/A', 'N/A'

def verifica_acao(ticker):
    for cell in pag_1["B"]:
        if cell.value == ticker:
            return True
    return False

def adicionar(nome, cotacao_str, dividendos, pvps, pls, div_ebitdas, data_hoje):
    linha_valor = None
    for cell in pag_1["B"]:
        if cell.value == nome:
            linha_valor = cell.row
            break
    
    if not linha_valor: return False

    try:
        preco_teto = pag_1.cell(row=linha_valor, column=3).value
        cotacao_float = float(cotacao_str.replace(',', '.'))
        preco_teto_float = float(str(preco_teto).replace('R$', '').replace(',', '.')) if preco_teto else 0.0
        diferenca = cotacao_float - preco_teto_float
        
        pag_1.cell(row=linha_valor, column=4).value = f"R${cotacao_float:.2f}"
        cell_dif = pag_1.cell(row=linha_valor, column=5)
        cell_dif.value = f"R${diferenca:.2f}/{data_hoje}"
        cell_dif.font = Font(color='008000' if diferenca > 0 else 'FF0000')
        
        pag_1.cell(row=linha_valor, column=6).value = pvps
        pag_1.cell(row=linha_valor, column=7).value = pls
        pag_1.cell(row=linha_valor, column=8).value = div_ebitdas
        pag_1.cell(row=linha_valor, column=9).value = dividendos
        return True
    except:
        return False

# --- FUNÇÕES DE INTERFACE ---

def text_bucar(label):
    sleep(0.5)
    texto = 'Olá, digite as ações desejadas abaixo.'
    for p in texto:
        if not app.winfo_exists(): break
        label.configure(text=label.cget("text") + p, text_color='#a9a9a9')
        label.update()
        sleep(0.08)

def buscar():
    verificadas = 0
    acoes = campo_busca.get().upper().split()
    for acao in acoes:
        if verifica_acao(acao):
            cotacao, dy, pvp, pl, div_ebitda = obter_dados(acao)
            if cotacao != 'N/A':
                if adicionar(acao, cotacao, dy, pvp, pl, div_ebitda, date.today().strftime("%d-%m")):
                    resultado.configure(text=f'A ação {acao} foi atualizada na tabela.\n')
                    verificadas += 1
        else:
            resultado.configure(text=f'O código {acao} não foi encontrado.\n')
            sleep(1.5)
            
    if verificadas > 0:
        resultado.configure(text=f'Atualização concluída, deseja salvar as alterações?')
        busca.configure(text='Salvar', command=salvar_alteracoes)

def buscar_thread():
    threading.Thread(target=buscar, daemon=True).start()

def help():
    messagebox.showinfo("Ajuda", "Digite os tickers separados por espaço.\nExemplo: PETR4 VALE3 ITUB4")

def salvar_alteracoes():
    try:
        tabela_acoes.save(planilha)
        Popen(f'explorer "{os.path.normpath(planilha)}"', shell=True)
    except:
        messagebox.showerror("Erro", "Feche a planilha antes de salvar!")

# --- INICIALIZAÇÃO DA PLANILHA ---
tabela_acoes = load_workbook(planilha)
pag_1 = tabela_acoes["Plan1"]

# --- UI (CUSTOMTKINTER - VOLTANDO AO ORIGINAL) ---
set_appearance_mode('Dark') 
app = CTk()
app.title('Busca ações')
app.geometry('500x290')
app.configure(fg_color="#333333")
app.resizable(False, False)

# Mantendo os espaçadores originais do seu código
espaçadorh = CTkLabel(app, text=' ', width=20)
espaçadorh.grid(row=0, rowspan=2, column=1)

espaçadorv = CTkLabel(app, text='', width=200, height=80)
espaçadorv.grid(row=0, column=2)

introducao = CTkLabel(app, text='', font=('Arial', 16))
introducao.grid(row=2, column=2, pady=5)

campo_busca = CTkEntry(app, placeholder_text='Digite os tickers separados por espaço', placeholder_text_color='#a9a9a9', font=('Arial', 12), corner_radius=6, width=350, height=30, fg_color='#282727', border_color='#282727')
campo_busca.grid(row=3, column=2)

busca = CTkButton(app, text='Buscar', command=buscar_thread, width=60, height=30, fg_color='#282727', font=('Arial', 12), text_color="#a9a9a9")
busca.grid(row=3, column=4, pady=10, padx=10)

resultado = CTkLabel(app, text='', font=('Arial', 12))
resultado.grid(row=5, column=2, columnspan=2)

frame = CTkFrame(app, width=50, height=300, fg_color='#282727', corner_radius=0)
frame.grid(row=0, column=0, rowspan=15)

ajuda = CTkButton(app, text="?", command=help, width=20, height=30, fg_color="transparent", bg_color='#282727', border_width=0)
ajuda.grid(row=10, column=0)

app.after(100, text_bucar, introducao)
app.mainloop()